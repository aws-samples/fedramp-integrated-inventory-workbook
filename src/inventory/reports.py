# (c) 2019 Amazon Web Services, Inc. or its affiliates. All Rights Reserved.
# License:
# This sample code is made available under the MIT-0 license. See the LICENSE file.
from datetime import datetime
import logging
from pathlib import PurePath
import os, os.path
from typing import List
import boto3
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from inventory.mappers import InventoryData

_logger = logging.getLogger("inventory.reports")
_logger.setLevel(os.environ.get("LOG_LEVEL", logging.INFO))
_current_dir_name = os.path.dirname(__file__)
_workbook_template_file_name = os.path.join(_current_dir_name, "SSP-A13-FedRAMP-Integrated-Inventory-Workbook-Template.xlsx")
_workbook_output_file_path = PurePath("/tmp/SSP-A13-FedRAMP-Integrated-Inventory.xlsx")
DEFAULT_REPORT_WORKSHEET_FIRST_WRITEABLE_ROW_NUMBER = 3

class CreateReportCommandHandler():
    def _write_cell_if_value_provided(self, worksheet: Worksheet, column:int, row: int, value: str):
        if value:
            worksheet.cell(column=column, row=row, value=value)

    def execute(self, inventory: List[InventoryData]) -> str:
        workbook = load_workbook(_workbook_template_file_name)
        reportWorksheetName = os.environ.get("REPORT_WORKSHEET_NAME", "Inventory")
        reportWorksheet = workbook[reportWorksheetName]
        rowNumber: int = int(os.environ.get("REPORT_WORKSHEET_FIRST_WRITEABLE_ROW_NUMBER", DEFAULT_REPORT_WORKSHEET_FIRST_WRITEABLE_ROW_NUMBER))

        _logger.info(f"writing {len(inventory)} rows into worksheet {reportWorksheetName} starting at row {rowNumber}")

        for inventory_row in inventory:
            self._write_cell_if_value_provided(reportWorksheet, 1, rowNumber, inventory_row.unique_id)
            self._write_cell_if_value_provided(reportWorksheet, 2, rowNumber, inventory_row.ip_address)
            self._write_cell_if_value_provided(reportWorksheet, 3, rowNumber, inventory_row.is_virtual)
            self._write_cell_if_value_provided(reportWorksheet, 4, rowNumber, inventory_row.is_public)
            self._write_cell_if_value_provided(reportWorksheet, 5, rowNumber, inventory_row.dns_name)
            self._write_cell_if_value_provided(reportWorksheet, 7, rowNumber, inventory_row.mac_address)
            self._write_cell_if_value_provided(reportWorksheet, 8, rowNumber, inventory_row.authenticated_scan_planned)
            self._write_cell_if_value_provided(reportWorksheet, 9, rowNumber, inventory_row.baseline_config)
            self._write_cell_if_value_provided(reportWorksheet, 12, rowNumber, inventory_row.asset_type)
            self._write_cell_if_value_provided(reportWorksheet, 13, rowNumber, inventory_row.hardware_model)
            self._write_cell_if_value_provided(reportWorksheet, 15, rowNumber, inventory_row.software_vendor)
            self._write_cell_if_value_provided(reportWorksheet, 16, rowNumber, inventory_row.software_product_name)
            self._write_cell_if_value_provided(reportWorksheet, 18, rowNumber, inventory_row.iir_diagram_label)
            self._write_cell_if_value_provided(reportWorksheet, 21, rowNumber, inventory_row.network_id)
            self._write_cell_if_value_provided(reportWorksheet, 22, rowNumber, inventory_row.owner)
            self._write_cell_if_value_provided(reportWorksheet, 23, rowNumber, inventory_row.owner)

            rowNumber += 1

        workbook.save(_workbook_output_file_path)

        _logger.info(f"completed saving inventory into {_workbook_output_file_path}")

        return str(_workbook_output_file_path)

class DeliverReportCommandHandler():
    def __init__(self, s3_client=boto3.client('s3')):
        self._s3_client = s3_client

    def execute(self, report_file_name: str) -> str:
        target_path = os.environ["REPORT_TARGET_BUCKET_PATH"]
        target_bucket = os.environ["REPORT_TARGET_BUCKET_NAME"]
        report_s3_key = os.path.join(target_path, f"{_workbook_output_file_path.stem}-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.xlsx")

        _logger.info(f"uploading file '{report_file_name}' to bucket '{target_bucket}' with key '{report_s3_key}'")

        object_data = open(report_file_name, "rb")

        self._s3_client.put_object(Bucket=target_bucket, Key=report_s3_key, Body=object_data)

        _logger.info(f"completed file upload")

        return f"https://{target_bucket}.s3.amazonaws.com/{report_s3_key}"